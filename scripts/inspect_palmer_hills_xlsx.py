from __future__ import annotations

from pathlib import Path

import openpyxl


def main() -> None:
    workbook_path = Path(__file__).resolve().parents[1] / "docs" / "palmer_hills_nassau_skins_recreated.xlsx"
    if not workbook_path.exists():
        raise SystemExit(f"Missing workbook: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    print("sheets:", wb.sheetnames)

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== {name} === rows={ws.max_row} cols={ws.max_column}")
        max_cols = min(16, ws.max_column)
        printed = 0
        for r in range(1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, max_cols + 1)]
            if not any(v not in (None, "") for v in row):
                continue
            print(r, row)
            printed += 1
            if printed >= 20:
                break


if __name__ == "__main__":
    main()
