from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    workbook_path = Path("docs/palmer_hills_nassau_skins_recreated.xlsx")
    wb = load_workbook(workbook_path, data_only=True)

    # --- Nassau: recompute segment To-Par sums from "Team Best 2" sheet ---
    team_best = wb["Team Best 2"]

    rows: list[tuple[int, int, int, int, int]] = []
    for row_index in range(4, 22):
        hole = team_best.cell(row_index, 1).value
        if hole is None:
            continue

        par = team_best.cell(row_index, 2).value
        team_a_to_par = team_best.cell(row_index, 4).value
        team_b_to_par = team_best.cell(row_index, 7).value
        team_c_to_par = team_best.cell(row_index, 10).value

        rows.append(
            (
                int(hole),
                int(par),
                int(team_a_to_par),
                int(team_b_to_par),
                int(team_c_to_par),
            )
        )

    def sum_to_par(hole_rows: list[tuple[int, int, int, int, int]]) -> tuple[int, int, int]:
        return (
            sum(r[2] for r in hole_rows),
            sum(r[3] for r in hole_rows),
            sum(r[4] for r in hole_rows),
        )

    front_rows = [r for r in rows if 1 <= r[0] <= 9]
    back_rows = [r for r in rows if 10 <= r[0] <= 18]

    print("NASSAU CHECK")
    print("- Front to-par:", sum_to_par(front_rows))
    print("- Back  to-par:", sum_to_par(back_rows))
    print("- Total to-par:", sum_to_par(rows))

    # --- Skins: recompute unique low gross winners from "Scorecard" sheet ---
    scorecard = wb["Scorecard"]
    header = [scorecard.cell(3, col).value for col in range(1, 17)]
    players = header[4:]

    skins: list[tuple[int, int, int, str]] = []
    for row_index in range(4, 22):
        hole = scorecard.cell(row_index, 1).value
        if hole is None:
            continue

        par = int(scorecard.cell(row_index, 2).value)
        scores = [scorecard.cell(row_index, col).value for col in range(5, 17)]
        scores_int = [int(s) for s in scores]

        low = min(scores_int)
        winners = [players[i] for i, s in enumerate(scores_int) if s == low]
        if len(winners) == 1:
            skins.append((int(hole), par, low, str(winners[0])))

    print("\nSKINS CHECK")
    print("- Skin count:", len(skins))
    print("- Skin holes:", skins)


if __name__ == "__main__":
    main()
